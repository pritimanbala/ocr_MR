"""Excel workbook parser backed by openpyxl."""
from __future__ import annotations
from pathlib import Path
from engineering_di.models import BoundingBox, Document, Page, Table, TableCell, Token
from engineering_di.parsers.base_parser import BaseParser

class ExcelParser(BaseParser):
    supported_extensions = frozenset({".xlsx", ".xls"})
    parser_name = "openpyxl"

    def parse(self, path: str | Path) -> Document:
        path = self.validate_path(path)
        if path.suffix.lower() == ".xls":
            raise RuntimeError("Legacy .xls parsing requires conversion to .xlsx before ingestion; install a converter service such as LibreOffice headless.")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Missing dependency 'openpyxl'. Install requirements.txt.") from exc
        wb = load_workbook(path, data_only=False, read_only=False)
        pages: list[Page] = []
        sheet_metadata: list[dict[str, object]] = []
        for si, ws in enumerate(wb.worksheets, start=1):
            cells: list[TableCell] = []
            tokens: list[Token] = []
            table_cell_ids: list[str] = []
            merged = {str(rng): rng.coord for rng in ws.merged_cells.ranges}
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    r, c = cell.row - 1, cell.column - 1
                    bbox = BoundingBox(c * 90.0, r * 20.0, (c + 1) * 90.0, (r + 1) * 20.0)
                    cid = f"p{si}.sheetcell.r{r}.c{c}"
                    text = str(value)
                    table_cell_ids.append(cid)
                    cells.append(TableCell(id=cid, bbox=bbox, page_number=si, row=r, column=c, text=text))
                    tokens.append(Token(id=f"{cid}.token", text=text, bbox=bbox, page_number=si, source="ocr"))
            table = Table(id=f"p{si}.sheet", bbox=BoundingBox(0, 0, max(ws.max_column, 1) * 90.0, max(ws.max_row, 1) * 20.0), page_number=si, cell_ids=table_cell_ids)
            metadata = {
                "sheet_name": ws.title,
                "merged_cells": list(merged.values()),
                "hidden_rows": [idx for idx, dim in ws.row_dimensions.items() if dim.hidden],
                "hidden_columns": [key for key, dim in ws.column_dimensions.items() if dim.hidden],
            }
            sheet_metadata.append(metadata)
            pages.append(Page(number=si, bbox=table.bbox, rotation=0, tokens=tokens, tables=[table], cells=cells, text_blocks=[], layout_regions=[]))
        return Document(source_path=str(path), page_count=len(pages), metadata={"workbook": {"sheet_names": wb.sheetnames, "properties": str(wb.properties)}, "sheets": sheet_metadata}, pages=pages, parser=self.parser_name, schema_version="engineering.document.v1")
